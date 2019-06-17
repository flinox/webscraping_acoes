'''
Versão inicial: By: Fernando Lino
Data: 04/10/2018
'''
import datetime
from bs4 import BeautifulSoup
import requests
import openpyxl
import json
from shutil import copyfile

class Acao:

    __taxa_emolumentos = (0.004934 / 100)
    __taxa_liquidacao = (0.0275 / 100)
    __website = 'http://cotacoes.economia.uol.com.br/acao/cotacoes-historicas.html?codigo=%s.SA'

    def __init__(self,acao, data, qtde, valor, corretagem):
        self.__acao = acao
        self.__data = data
        self.__qtde = qtde
        self.__valor = valor
        self.__corretagem = corretagem
        self.__total_bruto = self.__calcula_total_bruto(qtde,valor)
        self.__emolumentos = self.__calcula_emolumentos(self.__total_bruto)
        self.__liquidacao = self.__calcula_liquidacao(self.__total_bruto)
        self.__total_liquido = self.__calcula_total_liquido(self.__total_bruto,corretagem,self.__emolumentos,self.__liquidacao)
        self.data_cotacao = None
        self.valor_cotacao = None
        self.getCotacaoAtual()        

    # Tornar os atributos apenas leitura, os valores serao assignados somente no momento de criacao da instancia
    @property
    def acao(self):
        return self.__acao

    @property
    def data(self):
        return self.__data

    @property
    def qtde(self):
        return self.__qtde

    @property
    def valor(self):
        return self.__valor

    @property
    def corretagem(self):
        return self.__corretagem

    @property
    def total_bruto(self):
       return self.__total_bruto

    @property
    def emolumentos(self):
       return self.__emolumentos

    @property
    def liquidacao(self):
       return self.__liquidacao

    @property
    def total_liquido(self):
       return self.__total_liquido


    # Funcoes Calculos
    def __calcula_emolumentos(self,valor_bruto):
        return round(valor_bruto * self.__taxa_emolumentos, 2)

    def __calcula_liquidacao(self,valor_bruto):
        return round(valor_bruto * self.__taxa_liquidacao, 2)

    def __calcula_total_bruto(self,qtde, valor):
        return round((valor * qtde), 2)

    def __calcula_total_liquido(self,valor_bruto, corretagem, valor_emolumentos, valor_liquidacao):
        return round(valor_bruto - (corretagem + valor_emolumentos + valor_liquidacao), 2)


    def viewAcao(self):
        print('Ação: '+ self.acao +'\n'+
              'Data: ' + self.data + '\n' +
              'Qtde: ' + str(self.qtde) + '\n' +
              'Valor: ' + str(self.valor) + '\n' +
              'Total Bruto: ' + str(self.total_bruto) + '\n' +
              'Emolumentos: ' + str(self.emolumentos) + '\n' +
              'Liquidacao: ' + str(self.liquidacao) + '\n' +
              'Corretagem: ' + str(self.corretagem) + '\n' +
              'Total Liquido: ' + str(self.total_liquido) + '\n' +
              'Cotacao Data: ' + str(self.data_cotacao) + '\n' +
              'Cotacao Valor: ' + str(self.valor_cotacao) + '\n')

    def getAcao(self):
        return {'data' : self.data,
                'qtde' : self.qtde,
                'valor' : self.valor,
                'emolumentos' : self.emolumentos,
                'liquidacao' : self.liquidacao,
                'corretagem' : self.corretagem,
                'total_bruto': self.total_bruto,
                'total_liquido' : self.total_liquido,
                'cotacao_data' : self.data_cotacao,
                'cotacao_valor' : self.valor_cotacao}

    def getCotacaoAtual(self):
        self.__website = str(self.__website % self.__acao)
        r = requests.get(self.__website)
        pagina = BeautifulSoup(r.content, 'html.parser')

        valor = pagina.find('td', {'class': 'ultima'})
        valor = round(float(str(valor.text.strip()).replace(",", ".")), 2)

        self.data_cotacao = datetime.date.today().strftime("%d/%m/%Y")
        self.valor_cotacao = valor
        r.close()

class Carteira:

    def __init__(self):
        self.acoes = {}
        self.cotacao = {}

    def addAcao(self,acao,dados):
        self.acoes[acao] = dados

    def getCarteira(self):
        return self.acoes


class Arquivo_xlsx:

    xfile = None
    nome_arquivo = None
    tab = 'Ações'
    nome_arquivo_version = None

    def __init__(self,arquivo):
        self.xfile = openpyxl.load_workbook(arquivo + '.xlsx')
        self.nome_arquivo = arquivo

    def getArquivo(self):
        return self.xfile

    def saveArquivo(self, version):
        self.nome_arquivo_version = self.nome_arquivo + '_' + version + '.xlsx'
        self.xfile.save(self.nome_arquivo_version)

    def updAcao(self,acao, dados_acao):

        sheet = self.xfile[self.tab]
        for i in range(1, 50):
            if (sheet.cell(row=i, column=1).value == acao) or (sheet.cell(row=i, column=1).value == acao + 'F'):

                # Sequencia de acordo com o layout da planilha ACOES.xlsx
                sheet.cell(row=i, column=2).value = dados_acao['data']
                sheet.cell(row=i, column=4).value = dados_acao['qtde']
                sheet.cell(row=i, column=5).value = dados_acao['valor']
                sheet.cell(row=i, column=7).value = dados_acao['corretagem']
                sheet.cell(row=i, column=12).value = dados_acao['cotacao_valor']
                sheet.cell(row=i, column=14).value = dados_acao['cotacao_data']                

    def moveArquivo(self,destino):
        copyfile(self.nome_arquivo_version, destino + r'\\' + self.nome_arquivo_version)